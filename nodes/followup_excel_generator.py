"""
Follow-up Excel Generator Node - Creates Excel files for follow-up analyses
Generates budget estimates, cost breakup, and manpower estimates Excel files.
"""
import os
from datetime import datetime
from typing import Dict, Any, List

from openpyxl import Workbook

from state import GraphState
from config import (
    OUTPUT_DIR,
    BUDGET_ESTIMATE_COLUMNS,
    COST_BREAKUP_COLUMNS,
    MANPOWER_ESTIMATE_COLUMNS
)
from utils.excel_helper import (
    create_workbook,
    format_excel_sheet,
    apply_data_formatting
)
from nodes.followup_handler import mark_followup_complete


# Column key mappings for each follow-up type
BUDGET_COLUMN_KEYS = {
    "Phase": "phase",
    "Activity": "activity",
    "Cost Head": "cost_head",
    "Budget Type": "budget_type",
    "Apx.Qty": "apx_qty",
    "Unit": "unit",
    "Rate": "rate",
    "Apx.Budget": "apx_budget",
    "Contingency %": "contingency_percent"
}

COST_BREAKUP_COLUMN_KEYS = {
    "Work Name": "work_name",
    "Description": "description",
    "Category": "category",
    "Apx.Qty": "apx_qty",
    "Unit": "unit",
    "Apx. Material Cost": "apx_material_cost",
    "Total Cost": "total_cost"
}

MANPOWER_COLUMN_KEYS = {
    "Activity": "activity",
    "Role": "role",
    "Skill Level": "skill_level",
    "No. of Workers": "no_of_workers",
    "Productivity (Unit/Day)": "productivity_unit_per_day",
    "Duration (Days)": "duration_days",
    "Man-Days": "man_days",
    "Apx. Daily Rate": "apx_daily_rate",
    "Apx. Cost": "apx_cost"
}


def get_item_value(item: Dict[str, Any], column: str, column_keys: Dict[str, str]) -> Any:
    """
    Get the value for a specific column from an item dictionary.
    
    Args:
        item: Item dictionary from LLM response
        column: Column header name
        column_keys: Mapping of column names to dictionary keys
        
    Returns:
        Value for the column
    """
    key = column_keys.get(column, column.lower().replace(" ", "_").replace(".", "").replace("%", "percent"))
    value = item.get(key, "")
    
    # Handle None values
    if value is None:
        return ""
    
    return value


def write_items_to_sheet(ws, items: List[Dict[str, Any]], columns: List[str], column_keys: Dict[str, str]) -> int:
    """
    Write items to the worksheet.
    
    Args:
        ws: Worksheet to write to
        items: List of item dictionaries
        columns: List of column headers
        column_keys: Mapping of column names to dictionary keys
        
    Returns:
        Number of rows written
    """
    row_idx = 2  # Start after header row
    
    for item in items:
        for col_idx, column in enumerate(columns, 1):
            value = get_item_value(item, column, column_keys)
            ws.cell(row=row_idx, column=col_idx, value=value)
        row_idx += 1
    
    return row_idx - 2  # Return number of data rows


def generate_budget_excel(state: GraphState, items: List[Dict[str, Any]]) -> str:
    """
    Generate budget estimate Excel file.
    
    Args:
        state: Current graph state
        items: Budget items from analyzer
        
    Returns:
        Path to generated Excel file
    """
    wb = create_workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create Budget Estimate sheet
    ws = wb.create_sheet("Budget Estimate")
    
    # Apply header formatting
    format_excel_sheet(ws, BUDGET_ESTIMATE_COLUMNS)
    
    # Write items
    rows_written = write_items_to_sheet(ws, items, BUDGET_ESTIMATE_COLUMNS, BUDGET_COLUMN_KEYS)
    
    # Apply data formatting
    if rows_written > 0:
        apply_data_formatting(ws, 2, rows_written + 1, len(BUDGET_ESTIMATE_COLUMNS))
    
    # Generate filename
    responses = state.get("questionnaire_responses", {})
    project_type = responses.get("project_type", "project").replace(" ", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"budget_estimate_{project_type}_{timestamp}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    
    # Save workbook
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(filepath)
    
    return filepath


def generate_cost_breakup_excel(state: GraphState, items: List[Dict[str, Any]]) -> str:
    """
    Generate cost breakup Excel file.
    
    Args:
        state: Current graph state
        items: Cost breakup items from analyzer
        
    Returns:
        Path to generated Excel file
    """
    wb = create_workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create Cost Breakup sheet
    ws = wb.create_sheet("Cost Breakup")
    
    # Apply header formatting
    format_excel_sheet(ws, COST_BREAKUP_COLUMNS)
    
    # Write items
    rows_written = write_items_to_sheet(ws, items, COST_BREAKUP_COLUMNS, COST_BREAKUP_COLUMN_KEYS)
    
    # Apply data formatting
    if rows_written > 0:
        apply_data_formatting(ws, 2, rows_written + 1, len(COST_BREAKUP_COLUMNS))
    
    # Generate filename
    responses = state.get("questionnaire_responses", {})
    project_type = responses.get("project_type", "project").replace(" ", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"cost_breakup_{project_type}_{timestamp}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    
    # Save workbook
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(filepath)
    
    return filepath


def generate_manpower_excel(state: GraphState, items: List[Dict[str, Any]]) -> str:
    """
    Generate manpower estimates Excel file.
    
    Args:
        state: Current graph state
        items: Manpower items from analyzer
        
    Returns:
        Path to generated Excel file
    """
    wb = create_workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create Manpower Estimates sheet
    ws = wb.create_sheet("Manpower Estimates")
    
    # Apply header formatting
    format_excel_sheet(ws, MANPOWER_ESTIMATE_COLUMNS)
    
    # Write items
    rows_written = write_items_to_sheet(ws, items, MANPOWER_ESTIMATE_COLUMNS, MANPOWER_COLUMN_KEYS)
    
    # Apply data formatting
    if rows_written > 0:
        apply_data_formatting(ws, 2, rows_written + 1, len(MANPOWER_ESTIMATE_COLUMNS))
    
    # Generate filename
    responses = state.get("questionnaire_responses", {})
    project_type = responses.get("project_type", "project").replace(" ", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"manpower_estimates_{project_type}_{timestamp}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    
    # Save workbook
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(filepath)
    
    return filepath


def followup_excel_generator_node(state: GraphState) -> GraphState:
    """
    Main follow-up Excel generator node that creates the appropriate Excel file.
    
    Routes to the specific generator based on the analysis type and then
    marks the follow-up as complete.
    
    Args:
        state: Current graph state with followup_analysis_result
        
    Returns:
        Updated state with excel path added to followup_excel_paths
    """
    analysis_result = state.get("followup_analysis_result", {})
    
    if not analysis_result:
        print("\n❌ No analysis result found for Excel generation")
        return state
    
    analysis_type = analysis_result.get("type", "")
    items = analysis_result.get("items", [])
    
    if not items:
        print(f"\n⚠️ No items to write for {analysis_type} analysis")
        # Still mark as complete so we don't get stuck
        state = mark_followup_complete(state)
        return state
    
    print(f"\n📑 Follow-up Excel Generator: Creating {analysis_type} Excel file...")
    
    try:
        # Route to appropriate generator
        if analysis_type == "budget":
            filepath = generate_budget_excel(state, items)
        elif analysis_type == "cost_breakup":
            filepath = generate_cost_breakup_excel(state, items)
        elif analysis_type == "manpower":
            filepath = generate_manpower_excel(state, items)
        else:
            print(f"   ❌ Unknown analysis type: {analysis_type}")
            state = mark_followup_complete(state)
            return state
        
        # Add to followup_excel_paths
        if state.get("followup_excel_paths") is None:
            state["followup_excel_paths"] = []
        state["followup_excel_paths"].append(filepath)
        
        print(f"   ✅ Excel file saved: {filepath}")
        print(f"   📊 Total items: {len(items)}")
        
    except Exception as e:
        print(f"   ❌ Excel generation error: {str(e)}")
    
    # Mark the follow-up as complete
    state = mark_followup_complete(state)
    
    return state
