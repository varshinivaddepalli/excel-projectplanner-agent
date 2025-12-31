"""
Excel Helper Utilities - Formatting and workbook creation helpers
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import List


def create_workbook() -> Workbook:
    """
    Create a new workbook with default settings.
    
    Returns:
        New Workbook instance
    """
    wb = Workbook()
    return wb


def format_excel_sheet(ws, columns: List[str]) -> None:
    """
    Apply formatting to the Excel worksheet.
    
    Args:
        ws: Worksheet to format
        columns: List of column headers
    """
    # Define styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header formatting
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set column widths based on header names
    column_widths = {
        "Phase Name": 20,
        "Activity Name": 25,
        "Task Name": 35,
        "Task Category": 15,
        "Task Duration": 15,
        "Task Start Date": 15,
        "Task End Date": 15,
        "Work Quantity": 15,
        "Work Rate": 12,
        "Work UOM": 12,
        "Task Description": 40,
        "Priority": 10,
        "Predecessor": 15,
        "Dependency Type": 15,
        "Successor": 15
    }
    
    for col_idx, header in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        width = column_widths.get(header, 15)
        ws.column_dimensions[col_letter].width = width
    
    # Freeze the header row
    ws.freeze_panes = "A2"


def apply_data_formatting(ws, start_row: int, end_row: int, num_columns: int) -> None:
    """
    Apply formatting to data rows.
    
    Args:
        ws: Worksheet to format
        start_row: Starting row number (usually 2)
        end_row: Ending row number
        num_columns: Number of columns
    """
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Alternate row colors for better readability
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(1, num_columns + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = cell_alignment
            cell.border = thin_border
            
            # Apply alternate row coloring
            if row_idx % 2 == 0:
                cell.fill = light_fill


def add_project_info_sheet(wb: Workbook, project_info: dict) -> None:
    """
    Add a project information sheet to the workbook.
    
    Args:
        wb: Workbook to add sheet to
        project_info: Dictionary with project information
    """
    ws = wb.create_sheet("Project Info", 0)
    
    # Define styles
    header_font = Font(bold=True, size=12)
    value_font = Font(size=11)
    
    # Add project information
    info_rows = [
        ("Project Location", project_info.get("location", "")),
        ("Project Type", project_info.get("project_type", "")),
        ("Plan Level", project_info.get("plan_level", "")),
        ("Start Date", project_info.get("start_date", "")),
    ]
    
    # Add scale details
    scale = project_info.get("project_scale", {})
    for key, value in scale.items():
        if value:
            formatted_key = key.replace("_", " ").title()
            info_rows.append((formatted_key, str(value)))
    
    for row_idx, (label, value) in enumerate(info_rows, 1):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = header_font
        
        value_cell = ws.cell(row=row_idx, column=2, value=value)
        value_cell.font = value_font
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40

